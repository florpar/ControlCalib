import pandas as pd
from openpyxl import load_workbook
import os
import matplotlib
matplotlib.use('Agg')  # Use 'Agg' backend to avoid issues with GUI
import matplotlib.pyplot as plt
plt.ioff()  # Turn off interactive mode

rptext="C:\\GammaControl\\controlgeneral.txt"

def get_detector_number(txt_file):
    with open(txt_file, 'r') as file:
        for c_line in file:
            if c_line.startswith('Detector'):
                # Extract the detector number
                det_line = c_line.split('#')[1]
                det_nuspa=det_line.split("ACQ")[0]
                det_num = det_line.split()[0]
                if det_num == '1':
                    return 7
                elif det_num == '0':
                    return 5
                else: return int(det_num)

def load_detector_config(conf_file):
    config_data = {}
    with open(conf_file, 'r') as file:
        for line_conf in file:
            conf_key, conf_value = line_conf.strip().split('=')
            config_data[conf_key.strip()] = conf_value.strip()
    # Extract relevant values
    out_file = config_data['output_file']
    out_dir = config_data['output_dir']
    copy_out_file = config_data['copy_output_file']
    back_file = config_data['backup_file']
    return out_file, out_dir, copy_out_file, back_file

def load_detector_pico(conf_file):
    PI_INFO = {}
    with open(conf_file, 'r') as file:
        for line_conf in file:
            # Look for the line starting with 'Peak_INFO'
            if line_conf.startswith('Peak_INFO'):
                continue  # Skip the 'Peak_INFO=' line itself
            if ':' in line_conf:
                # Extract energy and resolution details from each line
                energy, reso_info = line_conf.split(':', 1)
                energy = float(energy.strip())
                # Parse the FWHM and FWTM from the resolution info
                reso_info = reso_info.strip().strip('()')  # Remove parentheses
                resols = reso_info.split(',')
                FWHM = float(resols[0].split(':')[1].strip())
                FWTM = float(resols[1].split(':')[1].strip())
                # Store the parsed information in the PI_INFO dictionary
                PI_INFO[energy] = {'FWHM': FWHM, 'FWTM': FWTM}
    return PI_INFO

# Function to read and concatenate data from Excel
def read_and_concatenate(excel_file):
    wb = load_workbook(excel_file)
    df_list = []
    for phoja in wb.sheetnames:
        hoj = wb[phoja]
        df = pd.DataFrame(hoj.values)
        df.dropna(how='all', inplace=True)
        header_row = df.iloc[0]
        df = df[1:].reset_index(drop=True)
        df.columns = header_row
        df_list.append(df)
    return pd.concat(df_list, ignore_index=True)

def reset_and_drop_index(datos):
    data_reset = datos.reset_index(drop=True)
    return data_reset

def filter_and_drop(datos, pico):
    filtered_data = datos[datos['ENERGY'].astype(float) == pico].drop('ENERGY', axis=1)
    return filtered_data

def plot_graphs(datos, pico, datos_info, output_dir):
    datos['Fecha'] = pd.to_datetime(datos['Fecha'], format='%d/%m/%Y')
    year_datos = datos['Fecha'].dt.strftime('%Y')
    datos.sort_values('Fecha', inplace=True)
    datos['Fecha'] = datos['Fecha'].dt.strftime('%d/%m')  # Format as dd/mm
    colores = datos['Calib'].map({'ok': 'green', 'calib': 'red'})
    fig, axs = plt.subplots(3, 1, figsize=(10, 15))
    size_datos = 100
    for ax in axs:
        ax.set_facecolor(color=(218.0 / 255.0, 221 / 255.0, 227 / 255.0))
        ax.yaxis.grid(True)
    # Plot Centroid vs Fecha
    axs[0].scatter(datos['Fecha'], datos['CENTROID'], color=colores, s=size_datos)
    axs[0].plot(datos['Fecha'], datos['CENTROID'], color='blue', linestyle='solid')
    axs[0].axhline(pico - 0.3, color='yellow', linestyle='solid')
    axs[0].axhline(pico + 0.3, color='yellow', linestyle='solid')
    axs[0].axhline(pico, color='magenta', linestyle='solid')
    axs[0].set_title('Centroid vs Fecha para %s keV en %s' % (pico, year_datos.iloc[0]))
    axs[0].set_ylabel('Centroid')
    axs[0].set_xticklabels([])
    # Plot FWHM vs Fecha
    axs[1].scatter(datos['Fecha'], datos['FWHM'], color=colores, s=size_datos)
    axs[1].plot(datos['Fecha'], datos['FWHM'], color='blue', linestyle='solid')
    axs[1].axhline(datos_info['FWHM'], color='yellow', linestyle='solid')
    axs[1].axhline((datos_info['FWHM'] / 1.1), color='magenta', linestyle='solid')
    axs[1].set_title('FWHM vs Fecha para %s keV en %s' % (pico, year_datos.iloc[0]))
    axs[1].set_ylabel('FWHM')
    axs[1].set_xticklabels([])
    # Plot FWTM vs Fecha
    axs[2].scatter(datos['Fecha'], datos['FWTM'], color=colores, s=size_datos)
    axs[2].plot(datos['Fecha'], datos['FWTM'], color='blue', linestyle='solid')
    axs[2].axhline(datos_info['FWTM'], color='yellow', linestyle='solid')
    axs[2].axhline((datos_info['FWTM'] / 1.1), color='magenta', linestyle='solid')
    axs[2].set_title('FWTM vs Fecha para %s keV en %s' % (pico, year_datos.iloc[0]))
    axs[2].set_ylabel('FWTM')
    # Rotate the x-axis labels for readability
    for ax in axs:
        plt.sca(ax)
        plt.xticks(rotation=45)    
    plt.tight_layout()
    file_name = '%s.png' % str(pico)
    save_path = os.path.join(output_dir, file_name)
    plt.savefig(save_path)
    plt.close(fig)

def main():
    # Parse the detector number from the text file
    dn=get_detector_number(rptext)
    config_file = os.path.join("C:\\PrograminfoDet", "PathDet%d.txt" % dn)
    lim_file = os.path.join("C:\\PrograminfoDet", "LimDet%d.txt" % dn)
    output_file, output_dir, copy_output_file, backup_file= load_detector_config(config_file)
    PICO_INFO=load_detector_pico(lim_file)
    # Load and concatenate Excel data
    excel_data = read_and_concatenate(output_file)
    # Reset index and drop previous index column
    excel_reset = reset_and_drop_index(excel_data)
    # Create a dictionary for each pico's filtered data
    pico_dict = {}
    for pico in PICO_INFO.keys():
        filtered_data = filter_and_drop(excel_reset, pico)
        pico_dict[pico] = filtered_data    
    # Plot graphs for each pico
    for pico, df in pico_dict.items():
        control = df['Calib']
        if not df.empty and not control.any()== 'descalibrado':
            plot_graphs(df, pico, PICO_INFO[pico], output_dir)
    return pico_dict

if __name__ == "__main__":
    pico_dict = main()
