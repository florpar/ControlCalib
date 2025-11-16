import pandas as pd
import numpy as np
import os
import shutil
import ctypes
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

rptext="C:\\GammaControl\\controlgeneral.txt"
current_dat = datetime.today().strftime('%d/%m/%Y')
tol_centro = 0.3
tol_FW = 0.5

def get_detector_number(txt_file):
    with open(txt_file, 'r') as file:
        for c_line in file:
            if c_line.startswith('Detector'):
                # Extract the detector number
                det_line = c_line.split('#')[1]
                det_nuspa=det_line.split("ACQ")[0]
                det_num = det_line.split()[0]
                if det_num == '1':
                    return 7
                elif det_num == '0':
                    return 5
                else: return int(det_num)

def load_detector_config(conf_file):
    config_data = {}
    with open(conf_file, 'r') as file:
        for line_conf in file:
            conf_key, conf_value = line_conf.strip().split('=')
            config_data[conf_key.strip()] = conf_value.strip()
    # Extract relevant values
    out_file = config_data['output_file']
    out_dir = config_data['output_dir']
    copy_out_file = config_data['copy_output_file']
    back_file = config_data['backup_file']
    return out_file, out_dir, copy_out_file, back_file

def load_detector_pico(conf_file):
    PI_INFO = {}
    with open(conf_file, 'r') as file:
        for line_conf in file:
            # Look for the line starting with 'Peak_INFO'
            if line_conf.startswith('Peak_INFO'):
                continue  # Skip the 'Peak_INFO=' line itself
            if ':' in line_conf:
                # Extract energy and resolution details from each line
                energy, reso_info = line_conf.split(':', 1)
                energy = float(energy.strip())
                # Parse the FWHM and FWTM from the resolution info
                reso_info = reso_info.strip().strip('()')  # Remove parentheses
                resols = reso_info.split(',')
                FWHM = float(resols[0].split(':')[1].strip())
                FWTM = float(resols[1].split(':')[1].strip())
                # Store the parsed information in the PI_INFO dictionary
                PI_INFO[energy] = {'FWHM': FWHM, 'FWTM': FWTM}
    return PI_INFO

def norm_energy_name(val):
    try:
        return "{0:.2f}".format(float(val))
    except Exception:
        return str(val)

def dic_rango_centro(fpat, tolerance):
    df = pd.read_csv(fpat, skiprows=4, header=0, delimiter=' +', engine="python")
    df.drop(columns=['ROI#', 'RANGE(', 'keV)', 'GROSS', 'NET', '+/-', 'keV).1', 'Bq', '+/-.1'], inplace=True)
    df.rename(columns={'(' : 'ENERGY', 'FW(1/10)' : 'FWTM'}, inplace=True)
    df = df[df['LIBRARY'] == "Eu-152"].drop(columns=['LIBRARY']).reset_index(drop=True)
    df['Fecha'] = current_dat
    df['Calib'] = ''
    # Creating the new dataframe with selected columns
    df2 = df[['Fecha', 'CENTROID', 'FWHM', 'FWTM', 'Calib', 'ENERGY']]
    df_new = df2.copy()
    df_new['ENERGY'] = pd.to_numeric(df_new['ENERGY'], errors='coerce')
    # Flag to determine if any row is out of range
    descalibrado_flag = False
    # Loop through rows to check if any row fails the condition
    for idx, row in df_new.iterrows():
        lower_bound = row['ENERGY'] - tolerance
        upper_bound = row['ENERGY'] + tolerance   
        # Check if CENTROID is outside the tolerance bounds
        if not (lower_bound <= row['CENTROID'] <= upper_bound):
            descalibrado_flag = True
            break  # If one row is out of tolerance, no need to check further
    # Update the Calib column for all rows based on the flag
    if descalibrado_flag:
        df2['Calib'] = 'descalibrado'
    else:
        df2['Calib'] = 'ok'
    return df2

def get_rows(ofil):
    wb = load_workbook(ofil)
    datos_excel = []
    # Loop through each sheet
    for phoja in wb.sheetnames:
        hoj = wb[phoja]
        # Convert worksheet values to a DataFrame
        df = pd.DataFrame(hoj.values)
        if df.empty:
            last_row = 1
            calib_status = None
        else:
            # Assign the first row as header, and reindex DataFrame
            df.columns = df.iloc[0]  # Set first row as header
            df = df.drop(0).reset_index(drop=True)  # Drop the first row
            # Remove rows that are entirely None
            df.dropna(how='all', inplace=True)
            # Check if the 'Calib' column exists, and handle missing columns
            if 'Calib' in df.columns:
                calib_status = df.iloc[-1]['Calib']  # Get the 'Calib' status of the last row
            else:
                calib_status = 'vacio'
            # Find the first completely empty row (if any) for writing new data
            last_row = df.index[df.isnull().all(axis=1)].min()
            last_row = last_row + 1 if pd.notna(last_row) else len(df) + 1
        # Append data for each sheet to the list
        datos_excel.append({'ENERGY': phoja, 'Calib': calib_status, 'LastRow': last_row})
    return pd.DataFrame(datos_excel)

def generate_alert(alerts):
    messages = []
    for _, row in alerts.iterrows():
        if row['Alerta'] == 'CENTROID':
            alert_title = 'CALIBRAR y volver a correr el job'
            err_type = 16
        else:
            alert_title = 'Atencion'
            err_type = 48
        message = "En {} el {} fuera de rango: {}".format(row['Sheet'], row['Alerta'], row['Valor'])
        messages.append(message)
    # Combine all alert messages and display them
    alert_message = "\n".join(messages)
    ctypes.windll.user32.MessageBoxA(0, alert_message, alert_title, err_type)

# Appending data to existing worksheets without creating new ones
def append_to_worksheet(filname, df, tab_name, startrow=None, header=False):
    # Load the workbook and the worksheet
    workbook = load_workbook(filname)
    if tab_name in workbook.sheetnames:
        ws = workbook[tab_name]
    else:
        ws = workbook.create_sheet(title=tab_name)
        ws = workbook[tab_name]
    # If startrow is not provided, append data at the end
    if startrow is None:
        startrow = ws.max_row + 1
    # Write the header if startrow is 1
    if startrow == 1:
        for col_idx, column in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=column)
        startrow += 1  # Move to the next row to write the data
    # Append data row by row to the existing worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for c_idx, value in enumerate(row):
            ws.cell(row=startrow + r_idx, column=c_idx + 1, value=value)   
    # Save the workbook
    workbook.save(filname)

# Checking range for centroid values
def rango_centro(df, tolerance):
    df_new = df.copy()
    df_new['ENERGY'] = pd.to_numeric(df_new['ENERGY'], errors='coerce')
    alerts = []
    for index, row in df_new.iterrows():
        lower_bound = row['ENERGY'] - tolerance
        upper_bound = row['ENERGY'] + tolerance
        if not (lower_bound <= row['CENTROID'] <= upper_bound):
            alerts.append({'Sheet': row['ENERGY'], 'Alerta': 'CENTROID', 'Valor': row['CENTROID']})
    return pd.DataFrame(alerts)

# Checking FWHM and FWTM values
def check_fwhm_fwtm(df, pico_info):
    df_new = df.copy()
    df_new['ENERGY'] = pd.to_numeric(df_new['ENERGY'], errors='coerce')
    alerts = []
    for index, row in df_new.iterrows():
        energy = row['ENERGY']
        if energy in pico_info:
            fwhm_bound = pico_info[energy]['FWHM']
            fwtm_bound = pico_info[energy]['FWTM']
            if row['FWHM'] > fwhm_bound:
                alerts.append({'Sheet': row['ENERGY'], 'Alerta': 'FWHM', 'Valor': row['FWHM']})
            if row['FWTM'] > fwtm_bound:
                alerts.append({'Sheet': row['ENERGY'], 'Alerta': 'FWTM', 'Valor': row['FWTM']})
    return pd.DataFrame(alerts)

dn = get_detector_number(rptext)
config_file = os.path.join("C:\\PrograminfoDet", "PathDet%d.txt" % dn)
lim_file = os.path.join("C:\\PrograminfoDet", "LimDet%d.txt" % dn)
output_file, output_dir, copy_output_file, backup_file= load_detector_config(config_file)
PICO_INFO=load_detector_pico(lim_file)
dic_last_rows = get_rows(output_file)
LAST_ROWS = dic_last_rows.LastRow.values.tolist()
control = dic_last_rows.Calib.values.tolist()
data_cal = dic_rango_centro(rptext, tol_centro)
PICOS = data_cal.ENERGY.values.tolist()
centroid_alerts = rango_centro(data_cal, tol_centro)
fwhm_fwtm_alerts = check_fwhm_fwtm(data_cal, PICO_INFO)

# Ensure the output file exists
if not os.path.exists(output_file):
    shutil.copy2(backup_file, output_file)

try:
    for pico, ult_celda in zip(PICOS, LAST_ROWS):
        datapico = data_cal.loc[data_cal['ENERGY'] == pico]
        tab_name_norm = norm_energy_name(pico)
        if 'vacio' in control:
            append_to_worksheet(output_file, datapico, tab_name_norm, startrow=ult_celda, header=True)
        elif 'descalibrado' in control:
            datapicocalib = datapico.replace('ok', 'calib')
            append_to_worksheet(output_file, datapicocalib, tab_name_norm, startrow=ult_celda, header=False)
        else:
            append_to_worksheet(output_file, datapico, tab_name_norm, startrow=ult_celda+1, header=False)
    # Verify that the workbook was saved
    shutil.copyfile(output_file, copy_output_file)
    shutil.copyfile(output_file, backup_file)
    # Generate alerts
    if not centroid_alerts.empty:
        generate_alert(centroid_alerts)
    elif not fwhm_fwtm_alerts.empty:
        generate_alert(fwhm_fwtm_alerts)
    else:
        ctypes.windll.user32.MessageBoxA(0, b'Todos los valores en rango', b'Todo OK :)', 64)
except:
    ctypes.windll.user32.MessageBoxA(0, b"Excel Abierto", b"Cerra y proba de nuevo si sigue avisale a Flor", 16)
